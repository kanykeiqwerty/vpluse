import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime

def save_to_excel(companies_data=None, individuals_data=None, filename='bankruptcy_data.xlsx'):
    
    if not companies_data and not individuals_data:
        print("Нет данных для сохранения")
        return
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        
        if companies_data:
            for row in companies_data:
                row.pop('guid', None)
            df_companies = pd.DataFrame(companies_data)
            df_companies.to_excel(writer, sheet_name='Юридические лица', index=False)
            print(f" Юридические лица: {len(companies_data)} записей")
        
        if individuals_data:
            for row in individuals_data:
                row.pop('guid', None)
            df_individuals = pd.DataFrame(individuals_data)
            df_individuals.to_excel(writer, sheet_name='Физические лица', index=False)
            print(f" Физические лица: {len(individuals_data)} записей")

        for sheet_name in writer.sheets:
            sheet = writer.sheets[sheet_name]

            for col_idx, col_cells in enumerate(sheet.iter_cols(), start=1):
                max_length = 0
                col_letter = get_column_letter(col_idx)

                for row_idx, cell in enumerate(col_cells, start=1):
                    value = cell.value

                    # Сделать заголовок жирным и больше
                    if row_idx == 1:
                        cell.font = Font(bold=True, size=12)

                    # Формат даты
                    if isinstance(value, (datetime, pd.Timestamp)):
                        cell.number_format = 'DD.MM.YYYY'
                        value = value.strftime('%d.%m.%Y')

                    if value:
                        # Учёт длинного текста, даже с переносами
                        max_length = max(max_length, len(str(value)))

                # Автоширина колонок (чтобы текст был виден полностью)
                sheet.column_dimensions[col_letter].width = max_length + 5  # +5 для отступа

    print(f" Данные сохранены в {filename}")
